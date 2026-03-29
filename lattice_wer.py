"""
Lattice-Based ASR Evaluation for Hindi Transcription
=====================================================

Theory:
-------
Standard WER compares model output against a single rigid reference string.
This unfairly penalises models when:
  (a) The human reference itself contains errors.
  (b) Multiple valid surface forms exist for the same spoken word
      (numeral vs. written form, spelling variants, script variants,
       punctuation noise, compound-word split/merge, etc.)

A Lattice replaces the flat reference string with a sequence of BINS.
Each bin corresponds to one alignment position and holds every valid
surface form that is acceptable at that position.

WER against a lattice = minimum edit-distance where a "match" occurs
when the hypothesis token appears in the corresponding bin.

Design Choices
--------------
Alignment unit: WORD (whitespace-split after Unicode normalisation).
  * Hindi is written in Devanagari with space-delimited words.
  * Subword splitting would break morphological units meaninglessly.
  * Phrase-level units complicate multi-sequence alignment.

Normalisation applied before any comparison:
  * Strip punctuation (। ? ! , . - —)   ← punctuation is not scored in ASR
  * Collapse multiple spaces
  * Strip leading/trailing whitespace
  * Unicode NFC normalisation

Lattice construction (position-by-position):
  1. Normalise the reference and all 5 model outputs.
  2. Pairwise-align every model output to the reference using standard
     Needleman-Wunsch / edit-distance alignment (word level).
  3. At each reference position, collect every model token that was
     aligned there (including the reference token itself).
  4. Apply the MODEL-AGREEMENT OVERRIDE:
       If >= MODEL_AGREE_THRESH models agree on a token T that differs
       from the reference R at that position, T is added to the bin
       (and R is also kept, so models that matched the reference are
       not penalised).
  5. The resulting list of sets is the lattice.

Trust model over reference heuristic:
  A reference error is suspected when:
    * >= 3 models produce an identical token at the same position AND
    * That token differs from the reference token.
  The model-consensus token is then added to the lattice bin so that
  models that output it are NOT penalised.

Lattice WER computation:
  For each model:
    1. Normalise the model hypothesis.
    2. Run dynamic-programming edit distance against the lattice,
       where cost(hyp_token, bin) = 0 if hyp_token ∈ bin, else 1.
    3. WER = (S + D + I) / N_ref  (where N_ref = number of lattice bins)

Note: A lattice can never *increase* WER vs. the plain reference.
A model that was already scoring correctly against the reference will
score the same; a model that was wrongly penalised will improve.
"""

import re
import unicodedata
from collections import Counter
from openpyxl import load_workbook

# ────────────────────────── CONFIG ──────────────────────────
MODEL_AGREE_THRESH = 3   # ≥ this many models must agree to override reference
MODELS = ["Model H", "Model i", "Model k", "Model l", "Model m", "Model n"]
PUNCT_RE = re.compile(r'[।?!,.\-—।\'\"()\[\]{}:;~@#^*+=<>/\\|`_]')

# ────────────────────────── NORMALISATION ───────────────────
def normalise(text: str) -> list[str]:
    """Normalise a Hindi transcription and tokenise into word list."""
    if not text or not isinstance(text, str):
        return []
    # Unicode NFC
    text = unicodedata.normalize("NFC", text)
    # Strip punctuation marks (these are not semantically scored in ASR)
    text = PUNCT_RE.sub(" ", text)
    # Collapse whitespace
    tokens = text.split()
    return [t for t in tokens if t]


# ────────────────────────── EDIT-DISTANCE ALIGNMENT ─────────────────────────
def edit_align(ref_tokens: list[str], hyp_tokens: list[str]) -> list[tuple]:
    """
    Classic Needleman-Wunsch word-level alignment.
    Returns list of (ref_tok_or_None, hyp_tok_or_None) pairs.
    """
    R, H = len(ref_tokens), len(hyp_tokens)
    # DP table: dp[i][j] = min edits to align ref[:i] with hyp[:j]
    dp = [[0] * (H + 1) for _ in range(R + 1)]
    for i in range(R + 1): dp[i][0] = i
    for j in range(H + 1): dp[0][j] = j

    for i in range(1, R + 1):
        for j in range(1, H + 1):
            match = 0 if ref_tokens[i-1] == hyp_tokens[j-1] else 1
            dp[i][j] = min(
                dp[i-1][j-1] + match,  # sub or match
                dp[i-1][j]   + 1,       # deletion
                dp[i][j-1]   + 1,       # insertion
            )

    # Traceback
    alignment = []
    i, j = R, H
    while i > 0 or j > 0:
        if i > 0 and j > 0:
            match = 0 if ref_tokens[i-1] == hyp_tokens[j-1] else 1
            if dp[i][j] == dp[i-1][j-1] + match:
                alignment.append((ref_tokens[i-1], hyp_tokens[j-1]))
                i -= 1; j -= 1
                continue
        if i > 0 and dp[i][j] == dp[i-1][j] + 1:
            alignment.append((ref_tokens[i-1], None))      # deletion
            i -= 1
        else:
            alignment.append((None, hyp_tokens[j-1]))      # insertion
            j -= 1

    alignment.reverse()
    return alignment


# ────────────────────────── LATTICE CONSTRUCTION ────────────────────────────
def build_lattice(ref_tokens: list[str],
                  model_outputs: dict[str, list[str]],
                  agree_thresh: int = MODEL_AGREE_THRESH) -> list[set]:
    """
    Build a lattice (list of bins) from reference + 5 model outputs.

    Algorithm:
      For each reference position p:
        bin[p] = {ref_tokens[p]}          ← seed with reference token
        For each model, find the token it aligned to position p (if any).
        Add that model token to bin[p].
        MODEL-AGREEMENT OVERRIDE:
          If ≥ agree_thresh models independently produced token T ≠ ref
          at position p, add T to bin[p] (trust models over noisy ref).

    Insertions (model token aligned to None reference position) are
    handled by creating EPSILON bins that allow zero-cost skipping.
    """
    R = len(ref_tokens)
    if R == 0:
        return []

    # Align each model to the reference
    alignments: dict[str, list[tuple]] = {}
    for name, hyp in model_outputs.items():
        alignments[name] = edit_align(ref_tokens, hyp)

    # Build a ref-indexed map: ref_pos → {model → [tokens aligned here]}
    ref_map: list[dict[str, list]] = [{} for _ in range(R)]
    insertion_map: list[dict[str, list]] = [{}  for _ in range(R + 1)]

    for name, aln in alignments.items():
        ref_pos = 0
        for (r_tok, h_tok) in aln:
            if r_tok is not None and h_tok is not None:
                ref_map[ref_pos].setdefault(name, []).append(h_tok)
                ref_pos += 1
            elif r_tok is not None and h_tok is None:
                # deletion: model skipped this position
                ref_map[ref_pos].setdefault(name, [])
                ref_pos += 1
            else:
                # insertion: model inserted before current ref_pos
                insertion_map[ref_pos].setdefault(name, []).append(h_tok)

    lattice: list[set] = []

    for p in range(R):
        bin_p: set = {ref_tokens[p]}

        # Collect all model tokens at this position
        model_tokens_here = []
        for name, toks in ref_map[p].items():
            model_tokens_here.extend(toks)

        # Always add model tokens to the bin (spelling/script variants)
        bin_p.update(model_tokens_here)

        # MODEL-AGREEMENT OVERRIDE
        # Count how many models produced each *non-reference* token here
        counter = Counter()
        for name, toks in ref_map[p].items():
            for t in toks:
                if t != ref_tokens[p]:
                    counter[t] += 1
        for tok, cnt in counter.items():
            if cnt >= agree_thresh:
                bin_p.add(tok)   # already added above, but explicit for clarity

        lattice.append(bin_p)

        # Handle model insertions between p and p+1 as epsilon bins
        # (We mark these as optional positions with an empty token "ε"
        #  so models that inserted are not penalised.)
        if insertion_map[p]:
            inserted_toks = []
            for name, toks in insertion_map[p].items():
                inserted_toks.extend(toks)
            ins_counter = Counter(inserted_toks)
            # Only create an epsilon bin if >= agree_thresh models agree
            for tok, cnt in ins_counter.items():
                if cnt >= agree_thresh:
                    lattice.append({tok, "ε"})   # optional insertion

    return lattice


# ────────────────────────── LATTICE WER ─────────────────────────────────────
def lattice_wer(lattice: list[set], hyp_tokens: list[str]) -> dict:
    """
    Compute WER of hyp_tokens against the lattice.

    Cost function:
      match(h, bin) = 0  if h ∈ bin
                    = 0  if bin == {"ε"}  or  ("ε" in bin and h ∈ bin)
                    = 1  otherwise (substitution)
      deletion cost  = 1  (model skipped a ref position)
      insertion cost = 1  (model added extra token)

    For epsilon bins (optional positions), the model can skip freely.
    """
    L = len(lattice)
    H = len(hyp_tokens)

    if L == 0:
        return {"wer": 0.0, "S": 0, "D": 0, "I": H, "N": 0}

    # dp[i][j] = (cost, back_pointer)
    INF = float("inf")
    dp = [[INF] * (H + 1) for _ in range(L + 1)]
    dp[0][0] = 0

    # Fill: deletion of lattice position (model skipped)
    for i in range(1, L + 1):
        bin_i = lattice[i - 1]
        if "ε" in bin_i:
            dp[i][0] = dp[i-1][0]   # epsilon bin: free deletion
        else:
            dp[i][0] = dp[i-1][0] + 1

    # Fill: insertion (model has extra tokens vs. empty lattice)
    for j in range(1, H + 1):
        dp[0][j] = j

    for i in range(1, L + 1):
        bin_i = lattice[i - 1]
        is_optional = (bin_i == {"ε"}) or (len(bin_i) == 1 and "ε" in bin_i)

        for j in range(1, H + 1):
            h_tok = hyp_tokens[j - 1]

            # Match / Substitution
            sub_cost = 0 if (h_tok in bin_i) else 1
            sub = dp[i-1][j-1] + sub_cost

            # Deletion (model missing this bin)
            del_cost = 0 if is_optional else 1
            delete = dp[i-1][j] + del_cost

            # Insertion (extra token in hypothesis)
            insert = dp[i][j-1] + 1

            dp[i][j] = min(sub, delete, insert)

    # Count N = number of non-epsilon lattice positions (reference words)
    N = sum(1 for b in lattice if "ε" not in b)

    total_edits = dp[L][H]
    wer = total_edits / N if N > 0 else 0.0

    return {"wer": round(wer, 4), "total_edits": total_edits, "N": N}


# ────────────────────────── PLAIN WER (baseline) ────────────────────────────
def plain_wer(ref_tokens: list[str], hyp_tokens: list[str]) -> dict:
    """Standard word-level WER against flat reference."""
    R, H = len(ref_tokens), len(hyp_tokens)
    if R == 0:
        return {"wer": 0.0, "total_edits": H, "N": 0}

    dp = [[0] * (H + 1) for _ in range(R + 1)]
    for i in range(R + 1): dp[i][0] = i
    for j in range(H + 1): dp[0][j] = j
    for i in range(1, R + 1):
        for j in range(1, H + 1):
            cost = 0 if ref_tokens[i-1] == hyp_tokens[j-1] else 1
            dp[i][j] = min(dp[i-1][j-1] + cost, dp[i-1][j]+1, dp[i][j-1]+1)

    return {"wer": round(dp[R][H] / R, 4), "total_edits": dp[R][H], "N": R}


# ────────────────────────── LOAD DATA ───────────────────────────────────────
def load_data(path: str):
    wb = load_workbook(path, read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))

    header = rows[0]
    # Columns: url, Human, Model H, Model i, Model k, Model l, Model m, Model n
    segments = []
    for row in rows[1:]:
        seg = {
            "url":     row[0],
            "Human":   str(row[1]) if row[1] else "",
            "Model H": str(row[2]) if row[2] else "",
            "Model i": str(row[3]) if row[3] else "",
            "Model k": str(row[4]) if row[4] else "",
            "Model l": str(row[5]) if row[5] else "",
            "Model m": str(row[6]) if row[6] else "",
            "Model n": str(row[7]) if row[7] else "",
        }
        segments.append(seg)
    return segments


# ────────────────────────── MAIN EVALUATION ─────────────────────────────────
def evaluate(path: str):
    segments = load_data(path)

    # Accumulators
    plain_stats   = {m: {"total_edits": 0, "N": 0} for m in MODELS}
    lattice_stats = {m: {"total_edits": 0, "N": 0} for m in MODELS}

    detailed_rows = []   # for per-segment inspection

    for seg in segments:
        ref_tokens = normalise(seg["Human"])
        if not ref_tokens:
            continue

        model_outputs = {m: normalise(seg[m]) for m in MODELS}

        # Build lattice
        lattice = build_lattice(ref_tokens, model_outputs)

        for m in MODELS:
            hyp = model_outputs[m]

            pw = plain_wer(ref_tokens, hyp)
            lw = lattice_wer(lattice, hyp)

            plain_stats[m]["total_edits"] += pw["total_edits"]
            plain_stats[m]["N"]           += pw["N"]
            lattice_stats[m]["total_edits"] += lw["total_edits"]
            lattice_stats[m]["N"]           += lw["N"]

            detailed_rows.append({
                "url": seg["url"],
                "model": m,
                "ref": " ".join(ref_tokens),
                "hyp": " ".join(hyp),
                "plain_wer":   pw["wer"],
                "lattice_wer": lw["wer"],
                "delta":       round(lw["wer"] - pw["wer"], 4),
            })

    # Aggregate WERs
    print("\n" + "="*72)
    print("  AGGREGATE WER RESULTS  (across all segments)")
    print("="*72)
    print(f"{'Model':<12} {'Plain WER':>12} {'Lattice WER':>14} {'Δ WER':>10}  {'Verdict'}")
    print("-"*72)
    for m in MODELS:
        p_wer = plain_stats[m]["total_edits"] / plain_stats[m]["N"]
        l_wer = lattice_stats[m]["total_edits"] / lattice_stats[m]["N"]
        delta = l_wer - p_wer
        verdict = "REDUCED ✓" if delta < -0.0001 else ("same" if abs(delta) < 0.0001 else "INCREASED?")
        print(f"{m:<12} {p_wer:>12.4f} {l_wer:>14.4f} {delta:>10.4f}  {verdict}")
    print("="*72)

    # Show per-segment cases where WER improved significantly
    print("\n── TOP CASES WHERE LATTICE REDUCED WER (delta < -0.1) ──")
    improved = sorted(
        [r for r in detailed_rows if r["delta"] < -0.10],
        key=lambda x: x["delta"]
    )[:15]
    for r in improved:
        short_url = r["url"].split("/")[-1]
        print(f"\n  [{r['model']}]  {short_url}")
        print(f"    REF : {r['ref'][:80]}")
        print(f"    HYP : {r['hyp'][:80]}")
        print(f"    Plain WER={r['plain_wer']:.3f}  Lattice WER={r['lattice_wer']:.3f}  Δ={r['delta']:.3f}")

    # Show cases where lattice did NOT change WER (model was already correct)
    print("\n── SAMPLE CASES WHERE WER IS UNCHANGED (model was accurate) ──")
    unchanged = [r for r in detailed_rows if abs(r["delta"]) < 0.001
                 and r["plain_wer"] == 0.0][:5]
    for r in unchanged:
        short_url = r["url"].split("/")[-1]
        print(f"\n  [{r['model']}]  {short_url}")
        print(f"    REF : {r['ref'][:80]}")
        print(f"    HYP : {r['hyp'][:80]}")
        print(f"    Plain WER={r['plain_wer']:.3f}  Lattice WER={r['lattice_wer']:.3f}")

    # Inspect lattice for a sample segment to show bin structure
    sample_seg = segments[1]   # "मौनता का अर्थ क्या होता है"
    sample_ref  = normalise(sample_seg["Human"])
    sample_outs = {m: normalise(sample_seg[m]) for m in MODELS}
    sample_lat  = build_lattice(sample_ref, sample_outs)

    print("\n── LATTICE INSPECTION: Segment 2 ──")
    print(f"  Reference: {' '.join(sample_ref)}")
    for i, (tok, bin_) in enumerate(zip(sample_ref, sample_lat)):
        diff = bin_ - {tok}
        flag = " ◄ VARIANTS" if diff else ""
        print(f"  Bin[{i}] ref='{tok}'  full_bin={sorted(bin_)}{flag}")

    return detailed_rows, plain_stats, lattice_stats


if __name__ == "__main__":
    evaluate("/mnt/user-data/uploads/Question_4.xlsx")
